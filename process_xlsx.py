import pandas as pd
from openpyxl import load_workbook

# 加载文件
workbook = load_workbook('Network VT3.xlsx')

# 获取所有表名
sheet_names = workbook.sheetnames
sheet_names

# 直接使用pandas读取该工作表的数据
df = pd.DataFrame(workbook['Sheet1'].values)

# 查看数据的基本信息
print('数据基本信息：')
df.info()

# 查看数据集行数和列数
rows, columns = df.shape

if rows < 100 and columns < 20:
    # 短表数据（行数少于100且列数少于20）查看全量数据信息
    print('数据全部内容信息：')
    print(df.to_csv(sep='\t', na_rep='nan'))
else:
    # 长表数据查看数据前几行信息
    print('数据前几行内容信息：')
    print(df.head().to_csv(sep='\t', na_rep='nan'))

# 获取第1行(index=0)作为列名
new_header = df.iloc[0]

# 从第2行(index=1)开始加载数据
df = df[1:]

# 设置新的列名
df.columns = new_header

# 重置索引
df = df.reset_index(drop=True)

# 查看数据的基本信息
print('数据基本信息：')
df.info()

# 查看数据集行数和列数
rows, columns = df.shape

if rows < 100 and columns < 20:
    # 短表数据（行数少于100且列数少于20）查看全量数据信息
    print('数据全部内容信息：')
    print(df.to_csv(sep='\t', na_rep='nan'))
else:
    # 长表数据查看数据前几行信息
    print('数据前几行内容信息：')
    print(df.head().to_csv(sep='\t', na_rep='nan'))

# 用下划线分割Link Name列，取第一个内容
df['SECTION'] = df['Link Name'].str.split('_').str[0]

# 将结果保存为新的 Excel 文件
new_file_path = 'Network_VT3_modified.xlsx'
df.to_excel(new_file_path, index=False)